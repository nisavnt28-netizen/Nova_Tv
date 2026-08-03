from flask import Flask, request, jsonify
import csv
import os
from dotenv import load_dotenv
from supabase import create_client
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

load_dotenv()

app = Flask(__name__)

UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Phase 3: field-selection configs now live in Supabase (table: uploads)
# instead of local JSON files, so they survive a Render restart/redeploy.
SUPABASE_URL = os.environ.get('SUPABASE_URL')
SUPABASE_KEY = os.environ.get('SUPABASE_KEY')
supabase = create_client(SUPABASE_URL, SUPABASE_KEY) if SUPABASE_URL and SUPABASE_KEY else None

# Target: uploads up to 2GB. Werkzeug streams the file straight to disk
# during save(), so this limit alone does not load anything into RAM.
app.config['MAX_CONTENT_LENGTH'] = 2 * 1024 * 1024 * 1024

ALLOWED_EXTENSIONS = {'csv', 'xlsx'}
PREVIEW_ROWS = 10  # only ever read a handful of rows for preview, never the whole file


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def parse_csv_preview(filepath):
    """Reads only the header + first PREVIEW_ROWS rows. Safe on a 2GB file
    because csv.reader is a row-by-row iterator, not a full-file load."""
    with open(filepath, newline='', encoding='utf-8', errors='replace') as f:
        reader = csv.reader(f)
        header = next(reader)
        rows = []
        for i, row in enumerate(reader):
            if i >= PREVIEW_ROWS:
                break
            rows.append(row)
    return header, rows


def parse_xlsx_preview(filepath):
    """read_only=True makes openpyxl stream rows instead of loading the
    whole workbook into memory -- same reasoning as the CSV path above."""
    wb = load_workbook(filepath, read_only=True)
    sheet = wb.active
    row_iter = sheet.iter_rows(values_only=True)
    header = list(next(row_iter))
    rows = []
    for i, row in enumerate(row_iter):
        if i >= PREVIEW_ROWS:
            break
        rows.append(list(row))
    wb.close()
    return header, rows


@app.route('/')
def home():
    return jsonify({'status': 'ok', 'message': 'db-to-api panel backend is running'})


@app.route('/upload', methods=['POST'])
def upload():
    if 'file' not in request.files:
        return jsonify({'error': 'send the file under form field name "file"'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'no file selected'}), 400

    if not allowed_file(file.filename):
        return jsonify({'error': 'only .csv and .xlsx are supported right now'}), 400

    filename = secure_filename(file.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)  # streamed to disk, not held in RAM

    ext = filename.rsplit('.', 1)[1].lower()
    try:
        if ext == 'csv':
            header, rows = parse_csv_preview(filepath)
        else:
            header, rows = parse_xlsx_preview(filepath)
    except Exception as e:
        return jsonify({'error': f'could not parse file: {str(e)}'}), 400

    return jsonify({
        'filename': filename,
        'columns': header,
        'preview_rows': rows,
        'note': f'showing first {len(rows)} rows only, not the full file'
    })


@app.route('/select-fields', methods=['POST'])
def select_fields():
    data = request.get_json(silent=True)
    if not data:
        return jsonify({'error': 'send a JSON body with "filename" and "selected_fields"'}), 400

    filename = data.get('filename')
    selected_fields = data.get('selected_fields')

    if not filename:
        return jsonify({'error': '"filename" is required'}), 400
    if not selected_fields or not isinstance(selected_fields, list):
        return jsonify({'error': '"selected_fields" must be a non-empty list of column names'}), 400

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if not os.path.exists(filepath):
        return jsonify({'error': f'no uploaded file found for "{filename}" -- upload it first'}), 404

    # Re-read the real columns from disk rather than trusting the client,
    # so a stale or made-up field name can't sneak into the config.
    ext = filename.rsplit('.', 1)[1].lower()
    try:
        if ext == 'csv':
            header, _ = parse_csv_preview(filepath)
        else:
            header, _ = parse_xlsx_preview(filepath)
    except Exception as e:
        return jsonify({'error': f'could not re-read file: {str(e)}'}), 400

    invalid = [f for f in selected_fields if f not in header]
    if invalid:
        return jsonify({
            'error': f'these fields do not exist in the file: {invalid}',
            'available_columns': header
        }), 400

    config = {
        'filename': filename,
        'all_columns': header,
        'selected_fields': selected_fields
    }

    if supabase is None:
        return jsonify({'error': 'Supabase not configured -- set SUPABASE_URL and SUPABASE_KEY'}), 500

    supabase.table('uploads').upsert(config).execute()

    return jsonify({'message': 'field selection saved', 'config': config})


@app.route('/config/<filename>', methods=['GET'])
def get_config(filename):
    if supabase is None:
        return jsonify({'error': 'Supabase not configured -- set SUPABASE_URL and SUPABASE_KEY'}), 500

    result = supabase.table('uploads').select('*').eq('filename', filename).execute()
    if not result.data:
        return jsonify({'error': f'no saved field-selection for "{filename}" yet -- call /select-fields first'}), 404

    return jsonify(result.data[0])


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
